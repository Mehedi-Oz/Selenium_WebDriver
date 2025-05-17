from selenium import webdriver
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
import time

workbook = load_workbook(
    "C:/GITHUB/Selenium_WebDriver/Basics_of_selenium/authentication.xlsx"
)
sheet = workbook.active

browser = webdriver.Chrome()
browser.maximize_window()

for row in sheet.iter_rows(min_row=6, max_row=sheet.max_row, values_only=True):
    username = row[0]
    password = row[1]

    browser.get("https://www.saucedemo.com/")
    time.sleep(1)
    browser.find_element(By.ID, "user-name").send_keys(username)
    browser.find_element(By.ID, "password").send_keys(password)
    browser.find_element(By.ID, "login-button").click()
    time.sleep(1)

    browser.find_element(By.ID, "react-burger-menu-btn").click()
    time.sleep(1)
    browser.find_element(By.ID, "logout_sidebar_link").click()
    time.sleep(1)

browser.quit()
